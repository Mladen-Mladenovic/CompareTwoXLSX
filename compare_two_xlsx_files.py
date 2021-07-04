#! python3
'''
	Takes two .xlsx docs and 
	compares the values for chosen columns.
	The program makes changes directly in the documents.
	THE TWO DOCS ARE DELETED AFTER THE ANALISYS.

	The keys in both documents have to correspond, 
	meaning the id has to match for the rows with the same data.

	The constants here (adresses and column letters) are saved in a file varDepo.py
'''


import os, send2trash, openpyxl, pprint ,varDepo, re, logging
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logging.disable(level=logging.CRITICAL)

# Find addresses for the two xlsx docs based on first part of their names which is always constant.
def get_files_path():
	absenceRe = re.compile('^Allowance.*') # Create a regex expression using fixed part of the doc name.
	googleRe = re.compile('^HR.*')
	workDir = os.getcwd() # Find out current working directory path.
	li = os.listdir() # Get list of names of the files in the current working directory.
	# Iter the file names to find the ones matching to the regex pattern.
	for i in li:
		if absenceRe.search(i) is not None:
			absenceDoc = i
		elif googleRe.search(i) is not None:
			googleDoc = i
		else:
			continue
	# Create full path by adding file name to the directory path.
	absencePath = os.path.join(workDir, absenceDoc)
	googlePath = os.path.join(workDir, googleDoc)
	return googlePath, absencePath


# Load xlsx document, requires doc address and sheet name, returne workbook and worksheet as a tuple.
def load_data(address, sheet):
	wb = openpyxl.load_workbook(address, data_only=True)
	if sheet == None:
		ws = wb.active
	else:
		ws = wb[sheet]
	logging.debug(f'ws= {ws}')
	return wb, ws

# Count number of cells until the first blank cell, return the integer
def maximum_row(ws, key_column):
	if type(key_column) == int:
		key_column = get_column_letter(key_column)
	count=0
	for cell in ws [key_column]:
		if cell.value == None: # Stop with the first blank cell
			break
		else:
			count +=1
	return count

# Remove diacritics, strip white space and remove interpunction, changes are made directly in the document.
def format_names(workSheet, col_num):
	ws=workSheet
	maxim = maximum_row (ws, col_num)
	logging.debug(f'maxim ={maxim}')

	for col in ws.iter_cols(min_row=2, min_col=col_num, max_row=maxim, max_col=col_num):
		for cell in col:
			if cell.value == None: # Stop with the first blank cell
				break
			else:
				cell.value = cell.value.strip()
				cell.value = cell.value.translate(str.maketrans({'Č':'C','č':'c','Ć':'C','ć':'c','Š':'S','š':'s','Đ':'Dj','đ':'dj','Ž':'Z','ž':'z','.':''}))

# Create a list of two dictionaries, one for each dataset, return a list.
def create_dict_list(ws, key_column ,value_column):
	first = {}
	second = {}
	col_list = [first, second]
	for i in range (len(value_column)):
		count=1
		for cell in ws[key_column]:
			if cell.value == None: # Stop with the first blank cell
				break
			else:
				index = value_column[i]+str(count) # Create cell name by combining letter from the value column with the number from the >> i << counter
				col_list[i][cell.value] = ws[index].value # Add value to the dictionary from the cell on the position saved in >> index << 
				count+=1
		col_list.append # Append the dictionary for a single dataset to the output-ready list
	logging.debug(f'col_list = \n{col_list}')
	return col_list

# Combine data from first name and last name columns, and fills column 'A' with the full name data.
def construct_full_name(ws):
	maxim= ws.max_row
	firstNames = []
	lastNames = []
	# Extract first name
	for col in ws.iter_cols(min_row=2, min_col=1, max_row=maxim, max_col=1):
		for cell in col:
			firstNames.append(cell.value)
	# Extract last name
	for col in ws.iter_cols(min_row=2, min_col=2, max_row=maxim, max_col=2):
		for cell in col:
			lastNames.append(cell.value)
	fullNames = []
	# Combine first and last name to list of full names
	for i in range(len(firstNames)):
		logging.debug(f'\nfirstname = {firstNames[i]}\nlastNames = {lastNames[i]}')
		fullNames.append(firstNames[i]+' '+lastNames[i])
	# Write full names in row A
	c= 0 #cell counter
	for col in ws.iter_cols(min_row=2, min_col=1, max_row=maxim, max_col=1):
		for cell in col:
			cell.value = fullNames[c]
			logging.debug(f'cell value {cell.value}')
			c+=1

# Generate a list of keys based on the first document.
def get_key_list (first_list):
	key_list = []
	for keys in first_list[0]:
		key_list.append(keys)
	return key_list

# Find differences in the two datasets.
def compare_columns(first_list, second_list, key_list):
	with open('outputFile.txt','w') as outputFile: # Initiate the output file
		c=1 # Numeration
		for v in range(2): # Two datasets to compare
			for i in key_list: # The keys should be the same in both datasets
				try:
					if first_list[v][i] != second_list[v][i]:
						if v+1 == 1:
							colName = 'Residual'
						else:
							colName = 'Vacation days'
						print(f'{c}) Difference found:')
						outputFile.write('Difference found:\n')
						print(f'Values for key >> {i} << in column {colName} do not correspond.')
						outputFile.write(f'Values for key >> {i} << in column {colName} do not correspond.\n')
						print(f'\t{FIRST_LIST_DISPLAY_NAME} = {first_list[v][i]}\n\t{SECOND_LIST_DISPLAY_NAME} = {second_list[v][i]}\n')
						outputFile.write(f'\t{FIRST_LIST_DISPLAY_NAME} = {first_list[v][i]}\n\t{SECOND_LIST_DISPLAY_NAME} = {second_list[v][i]}\n\n')
						c+=1
				except KeyError:
					# Common KeyError if the datasets keys aren't the same
					print(f"\n{c}) DATA MISSING - The key >>> {i} <<< exists only in the first document.\nPlease ensure both documents have corresponding keys.\n")
					outputFile.write(f"\n{c}) The key >>> {i} <<< exists only in one document.\nPlease ensure both documents have corresponding keys.\n\n")
					c+=1


KEY_COLUMN = varDepo.variables[0]['KEY_COLUMN'] # Has to be same for both documents

# Both column letters need to be in the order of lookup.
FIRST_VALUE_COLUMN = varDepo.variables[0]['FIRST_VALUE_COLUMN']
SECOND_VALUE_COLUMN = varDepo.variables[0]['SECOND_VALUE_COLUMN']

FIRST_LIST_DISPLAY_NAME=varDepo.variables[0]['FIRST_LIST_DISPLAY_NAME']
SECOND_LIST_DISPLAY_NAME= varDepo.variables[0]['SECOND_LIST_DISPLAY_NAME']

# Get addresses for the two xlsx docs
google_doc,absence_doc = get_files_path()

# First doc
wb, ws = load_data(google_doc, 'Employees') # Load the document.
format_names(ws, 1)	# Fix input errors for keys and standardise format.
wb.save(google_doc) # Save shanges in the document.
first_list = create_dict_list(ws, KEY_COLUMN, FIRST_VALUE_COLUMN) # Extracts data in a list.

# Second doc
wb, ws = load_data(absence_doc, None) # Load the document.
construct_full_name(ws) # Form full name from 'First name' and 'Last name' columns.
format_names(ws, 1) # Fix input errors in keys and standardise format.
wb.save(absence_doc) # Save changes to the document.
second_list = create_dict_list(ws, KEY_COLUMN, SECOND_VALUE_COLUMN) # Extract data in a list.

# Analysis
key_list= get_key_list(first_list) # Generate a list of keys.
compare_columns(first_list, second_list, key_list) # Compare data sets and output findings.

# Delete xlsx docs.
send2trash.send2trash(google_doc)
send2trash.send2trash(absence_doc)